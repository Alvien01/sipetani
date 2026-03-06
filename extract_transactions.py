import openpyxl
import sys
from datetime import datetime

XLSX_PATH = r'c:\xampp\htdocs\sipetani\xReport2025.xlsx'
OUTPUT_PATH = r'c:\xampp\htdocs\sipetani\database\seeders\TransactionSeeder2025.php'

# Column mapping (0-indexed) for 18-col sheets:
# Col 1  = No Transaksi
# Col 4  = Tanggal (datetime)
# Col 8  = Jml Item (total_buy)
# Col 13 = Total Akhir (total_payment)

print("Loading workbook... (this may take a while)", flush=True)
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f"Total sheets: {len(sheets)}", flush=True)

transactions = []
seen_no = set()  # avoid duplicate transaction numbers

for idx, sheet_name in enumerate(sheets):
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 14:
            continue

        # Col 1 = No Transaksi (string like '000001/KSR/UTM/0123')
        no_trx = row[1]
        # Col 4 = Tanggal
        tanggal = row[4]
        # Col 8 = Jml Item
        jml_item = row[8]
        # Col 13 = Total Akhir
        total_akhir = row[13]

        if not isinstance(no_trx, str) or '/' not in no_trx:
            continue
        if no_trx in seen_no:
            continue
        if not isinstance(tanggal, datetime):
            continue
        if jml_item is None or total_akhir is None:
            continue

        try:
            qty = max(1, int(round(float(jml_item))))
            total = round(float(total_akhir), 2)
        except (TypeError, ValueError):
            continue

        if total <= 0:
            continue

        seen_no.add(no_trx)
        transactions.append({
            'date_sale': tanggal.date().strftime('%Y-%m-%d'),
            'total_buy': qty,
            'total_payment': total,
        })

    if (idx + 1) % 500 == 0:
        print(f"  Processed {idx+1}/{len(sheets)} sheets, extracted {len(transactions)} transactions", flush=True)

wb.close()
print(f"\nTotal extracted: {len(transactions)} transactions", flush=True)
print("Sample (first 5):")
for t in transactions[:5]:
    print(f"  {t}")
print("Sample (last 5):")
for t in transactions[-5:]:
    print(f"  {t}")

# ------- Generate PHP Seeder -------
print("\nGenerating TransactionSeeder.php...", flush=True)
now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
CHUNK_SIZE = 200

lines = []
lines.append("<?php\n")
lines.append("\nnamespace Database\\Seeders;\n")
lines.append("\nuse Illuminate\\Database\\Seeder;\n")
lines.append("use Illuminate\\Support\\Facades\\DB;\n")
lines.append("\nclass TransactionSeeder extends Seeder\n")
lines.append("{\n")
lines.append("    public function run(): void\n")
lines.append("    {\n")
lines.append("        // Get all product IDs from the database\n")
lines.append("        $productIds = DB::table('products')->pluck('id')->toArray();\n\n")
lines.append("        if (empty($productIds)) {\n")
lines.append("            $this->command->warn('No products found. Please seed products first.');\n")
lines.append("            return;\n")
lines.append("        }\n\n")
lines.append(f"        $this->command->info('Seeding {len(transactions)} transactions...');\n\n")
lines.append("        $data = [\n")

for t in transactions:
    lines.append(
        f"            ['product_id' => $productIds[array_rand($productIds)], "
        f"'date_sale' => '{t['date_sale']}', "
        f"'total_buy' => {t['total_buy']}, "
        f"'total_payment' => {t['total_payment']}, "
        f"'created_at' => '{now}', 'updated_at' => '{now}'],\n"
    )

lines.append("        ];\n\n")
lines.append(f"        foreach (array_chunk($data, {CHUNK_SIZE}) as $chunk) {{\n")
lines.append("            DB::table('transactions')->insert($chunk);\n")
lines.append("        }\n\n")
lines.append(f"        $this->command->info('Done! {len(transactions)} transactions seeded.');\n")
lines.append("    }\n")
lines.append("}\n")

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.writelines(lines)

print(f"Seeder written to: {OUTPUT_PATH}")
print("DONE!")
